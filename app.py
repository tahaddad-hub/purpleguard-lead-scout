import streamlit as st
import anthropic
import requests
import openpyxl
import ast
import re
import pandas as pd
from datetime import datetime
from supabase_client import get_supabase_client

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Sales Radar",
    page_icon="📡",
    layout="wide"
)

# ─────────────────────────────────────────────
# GLOBAL CSS — Card styling
# ─────────────────────────────────────────────
st.markdown("""
<style>
.sr-card {
    background: #FFFFFF;
    border: 1px solid #DDE4ED;
    border-radius: 10px;
    padding: 14px 16px 10px 16px;
    margin-bottom: 4px;
}
.card-name {
    font-size: 15px;
    font-weight: 600;
    color: #0F1923;
    margin-bottom: 8px;
    line-height: 1.4;
}
.sr-badge {
    display: inline-block;
    font-size: 11px;
    padding: 2px 8px;
    border-radius: 20px;
    background: #EEF2F7;
    color: #4A6080;
    border: 1px solid #DDE4ED;
    margin-right: 4px;
    margin-bottom: 4px;
}
.sr-badge-city {
    background: #EBF5FF;
    color: #1B6CA8;
    border-color: #BEDAF0;
}
.card-label {
    font-size: 10px;
    text-transform: uppercase;
    letter-spacing: 0.07em;
    color: #B0BCCC;
    font-weight: 600;
    margin-top: 8px;
    margin-bottom: 2px;
}
.card-vendors {
    font-size: 12px;
    color: #0F1923;
    font-weight: 500;
    margin-bottom: 4px;
}
.card-about {
    font-size: 12px;
    color: #4A6080;
    line-height: 1.5;
}
.card-divider {
    border: none;
    border-top: 1px solid #EEF2F7;
    margin: 8px 0;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
if "user" not in st.session_state:
    st.session_state.user = None
if "user_profile" not in st.session_state:
    st.session_state.user_profile = None
if "detected_country" not in st.session_state:
    st.session_state.detected_country = None
if "results" not in st.session_state:
    st.session_state.results = []
if "search_location" not in st.session_state:
    st.session_state.search_location = ""

# ─────────────────────────────────────────────
# LOGIN SCREEN
# ─────────────────────────────────────────────
def show_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("## 📡 Sales Radar")
        st.markdown("#### Please sign in to continue")
        st.divider()
        email    = st.text_input("Email", placeholder="your@email.com", autocomplete="off")
        password = st.text_input("Password", type="password", placeholder="••••••••", autocomplete="new-password")
        login_button = st.button("Sign In", type="primary", use_container_width=True)

        if login_button:
            if not email or not password:
                st.error("Please enter your email and password.")
                return
            try:
                supabase = get_supabase_client()
                response = supabase.auth.sign_in_with_password({"email": email, "password": password})
                if response.user:
                    profile = supabase.table("users").select("*").eq("id", response.user.id).single().execute()
                    if profile.data:
                        st.session_state.user         = response.user
                        st.session_state.user_profile = profile.data
                        st.rerun()
                    else:
                        st.error("User profile not found. Please contact your administrator.")
                else:
                    st.error("Invalid email or password.")
            except Exception:
                st.error("Login failed. Please check your credentials.")

# ─────────────────────────────────────────────
# LOGOUT
# ─────────────────────────────────────────────
def logout():
    try:
        supabase = get_supabase_client()
        supabase.auth.sign_out()
    except:
        pass
    st.session_state.user         = None
    st.session_state.user_profile = None
    st.session_state.results      = []
    st.rerun()

# ─────────────────────────────────────────────
# LOAD COUNTRIES AND CITIES
# ─────────────────────────────────────────────
@st.cache_data(ttl=3600)
def load_countries_and_cities():
    try:
        supabase = get_supabase_client()
        countries_response = supabase.table("countries").select("id, name").eq("is_active", True).order("name").execute()
        cities_response    = supabase.table("cities").select("name, country_id").eq("is_active", True).order("name").execute()
        country_map  = {c["id"]: c["name"] for c in countries_response.data}
        cities_dict  = {}
        for city in cities_response.data:
            country_name = country_map.get(city["country_id"])
            if country_name:
                if country_name not in cities_dict:
                    cities_dict[country_name] = []
                cities_dict[country_name].append(city["name"])
        for country in cities_dict:
            cities_dict[country] = sorted(cities_dict[country])
        return cities_dict
    except Exception as e:
        st.error("Could not load countries from database: " + str(e))
        return {}

# ─────────────────────────────────────────────
# LOAD INDUSTRIES — returns list of names + map
# ─────────────────────────────────────────────
@st.cache_data(ttl=3600)
def load_industries():
    try:
        supabase  = get_supabase_client()
        response  = supabase.table("industries").select("id, name").order("name").execute()
        names     = [i["name"] for i in response.data]
        id_map    = {i["name"]: i["id"] for i in response.data}
        return names, id_map
    except Exception:
        return [], {}

# ─────────────────────────────────────────────
# LOAD SPECIALTIES — filtered by industry_id
# ─────────────────────────────────────────────
@st.cache_data(ttl=3600)
def load_specialties(industry_id=None):
    try:
        supabase = get_supabase_client()
        query    = supabase.table("specialties").select("name").eq("is_active", True).order("name")
        if industry_id:
            query = query.eq("industry_id", industry_id)
        response = query.execute()
        return [s["name"] for s in response.data]
    except Exception:
        return []

# ─────────────────────────────────────────────
# LOAD TENANT OWN DOMAINS
# ─────────────────────────────────────────────
def load_own_domains(tenant_id):
    try:
        supabase = get_supabase_client()
        response = supabase.table("tenants").select("own_domains").eq("id", tenant_id).single().execute()
        if response.data and response.data.get("own_domains"):
            domains = [d.strip().lower() for d in response.data["own_domains"].split(",")]
            return [d for d in domains if d]
        return []
    except:
        return []

# ─────────────────────────────────────────────
# AUTO-DETECT USER COUNTRY
# ─────────────────────────────────────────────
def detect_user_country(available_countries):
    try:
        response = requests.get("https://ipapi.co/json/", timeout=3)
        data     = response.json()
        detected = data.get("country_name", "").strip()
        if detected in available_countries:
            return detected
        detected_lower = detected.lower()
        for country in available_countries:
            if detected_lower == country.lower():
                return country
        for country in available_countries:
            if detected_lower in country.lower() or country.lower() in detected_lower:
                return country
    except:
        pass
    return "Egypt" if "Egypt" in available_countries else available_countries[0]

# ─────────────────────────────────────────────
# BUILD COUNTRY LIST
# ─────────────────────────────────────────────
def build_country_list(cities_dict, user_country):
    all_countries = sorted(list(cities_dict.keys()))
    if user_country in all_countries:
        return [user_country] + [c for c in all_countries if c != user_country]
    return all_countries

# ─────────────────────────────────────────────
# WEB SEARCH
# ─────────────────────────────────────────────
def search_web(query, serper_key):
    url      = "https://google.serper.dev/search"
    headers  = {"X-API-KEY": serper_key, "Content-Type": "application/json"}
    payload  = {"q": query, "num": 10}
    response = requests.post(url, headers=headers, json=payload)
    results  = response.json()
    snippets = []
    url_map  = {}

    if "answerBox" in results:
        snippets.append("Summary: " + results["answerBox"].get("snippet", ""))
    if "organic" in results:
        for r in results["organic"][:10]:
            link    = r.get("link", "")
            title   = r.get("title", "")
            snippet = r.get("snippet", "")
            snippets.append("- " + title + ": " + snippet + " | URL: " + link)
            if link:
                try:
                    domain = link.split("//")[-1].split("/")[0].replace("www.", "")
                    if domain and domain not in url_map:
                        url_map[domain] = link
                except:
                    pass
    return "\n".join(snippets), url_map

# ─────────────────────────────────────────────
# EXTRACT DOMAIN
# ─────────────────────────────────────────────
def extract_domain(url):
    if not url:
        return ""
    try:
        return url.split("//")[-1].split("/")[0].replace("www.", "").lower().strip()
    except:
        return ""

# ─────────────────────────────────────────────
# DIRECTORY DOMAINS
# ─────────────────────────────────────────────
DIRECTORY_DOMAINS = [
    "clutch.co", "linkedin.com", "facebook.com", "twitter.com",
    "instagram.com", "yellowpages.com", "yelp.com", "bloomberg.com",
    "crunchbase.com", "zoominfo.com", "glassdoor.com", "indeed.com",
    "google.com", "wikipedia.org", "forbes.com", "reuters.com"
]

def is_directory_url(url):
    domain = extract_domain(url)
    return any(d in domain for d in DIRECTORY_DOMAINS)

# ─────────────────────────────────────────────
# PARSE AI RESPONSE
# ─────────────────────────────────────────────
def clean_and_parse(raw):
    raw = re.sub(r"```[a-z]*", "", raw).replace("```", "").strip()
    raw = raw.encode("ascii", "ignore").decode("ascii")
    raw = re.sub(r",\s*\]", "]", raw)
    raw = re.sub(r",\s*\}", "}", raw)
    return ast.literal_eval(raw)

# ─────────────────────────────────────────────
# SAVE TO SUSPECTS WITH DEDUPLICATION
# ─────────────────────────────────────────────
def save_to_suspects(df, country, objective, industry, specialty, company_size, user_profile):
    try:
        supabase       = get_supabase_client()
        own_domains    = load_own_domains(user_profile.get("tenant_id"))
        saved_count    = 0
        skipped_count  = 0
        excluded_count = 0

        for _, row in df.iterrows():
            company_name = row.get("Company Name", "").strip()
            website      = row.get("Website", "").strip()
            domain       = extract_domain(website)

            # EXCLUSION — Own company domains
            if domain and any(own in domain for own in own_domains):
                excluded_count += 1
                continue

            # Strip directory URLs
            if website and is_directory_url(website):
                website = ""
                domain  = ""

            # DEDUPLICATION — By domain
            if domain:
                existing = supabase.table("suspects").select("id").ilike("website", f"%{domain}%").execute()
                if existing.data and len(existing.data) > 0:
                    skipped_count += 1
                    continue

            # DEDUPLICATION — By company name
            if not domain and company_name:
                existing = supabase.table("suspects").select("id").ilike("company_name", company_name).execute()
                if existing.data and len(existing.data) > 0:
                    skipped_count += 1
                    continue

            record = {
                "company_name":  company_name,
                "city":          row.get("City", ""),
                "country":       country,
                "website":       website,
                "client_base":   row.get("Client Base", ""),
                "experience":    row.get("Experience", ""),
                "company_size":  row.get("Company Size", company_size),
                "known_vendors": row.get("Known Vendors", ""),
                "objective":     objective,
                "status":        "new",
                "is_active":     True,
                "phone":         "",
                "address":       "",
                "tenant_id":     user_profile.get("tenant_id"),
                "user_id":       user_profile.get("id"),
                "created_at":    datetime.utcnow().isoformat()
            }

            supabase.table("suspects").insert(record).execute()
            saved_count += 1

        return saved_count, skipped_count, excluded_count

    except Exception as e:
        st.error("Error saving to database: " + str(e))
        return 0, 0, 0

# ─────────────────────────────────────────────
# EXCLUDE COMPANY — Delete from suspects table
# ─────────────────────────────────────────────
def exclude_from_suspects(company_name, website, user_profile):
    try:
        supabase  = get_supabase_client()
        tenant_id = user_profile.get("tenant_id")
        domain    = extract_domain(website)

        if domain:
            supabase.table("suspects")\
                .delete()\
                .ilike("website", f"%{domain}%")\
                .eq("tenant_id", tenant_id)\
                .execute()
        elif company_name:
            supabase.table("suspects")\
                .delete()\
                .ilike("company_name", company_name)\
                .eq("tenant_id", tenant_id)\
                .execute()
    except Exception as e:
        st.error("Could not exclude company: " + str(e))

# ─────────────────────────────────────────────
# BUILD DYNAMIC SEARCH QUERIES — Zero hardcoding
# ─────────────────────────────────────────────
def build_search_queries(industry, specialty, company_size, location_for_search):
    ind  = industry  if industry  != "All Industries"  else ""
    spec = specialty if specialty != "All Specialties" else ""

    size_terms = {
        "Small (under 20)":       "small companies startups",
        "Medium (20-200)":        "medium size companies",
        "Large (200+)":           "large enterprise companies",
        "Any Size":               ""
    }
    size_term = size_terms.get(company_size, "")

    # Build base term: specialty is more specific than industry
    primary = spec if spec else ind
    secondary = ind if spec else ""

    queries = [
        f"{primary} companies {location_for_search}",
        f"{primary} {secondary} {location_for_search} {size_term}",
        f"list of {primary} companies {location_for_search}",
        f"{primary} businesses {location_for_search}",
        f"{primary} firms {location_for_search}",
        f"top {primary} companies {location_for_search} {size_term}",
    ]

    # Clean up double spaces
    queries = [" ".join(q.split()) for q in queries if q.strip()]
    return queries[:6]

# ─────────────────────────────────────────────
# BUILD DYNAMIC AI PROMPT — Zero hardcoding
# ─────────────────────────────────────────────
def build_prompt(industry, specialty, company_size, objective, location_for_search,
                 all_results, num_leads, tenant_name):

    ind  = industry  if industry  != "All Industries"  else ""
    spec = specialty if specialty != "All Specialties" else ""
    size = company_size if company_size != "Any Size" else ""

    limit_instruction = f"up to {num_leads}" if num_leads < 9999 else "as many as possible — do not limit yourself"

    prompt = (
        f"You are a strict business development researcher.\n"
        f"Find companies located in {location_for_search}.\n"
        f"Do NOT include companies from other locations.\n"
    )

    if tenant_name:
        prompt += f"IMPORTANT: Do NOT include '{tenant_name}' or any of its brands — this is our own company.\n"

    if ind:
        prompt += f"Industry: {ind}\n"
    if spec:
        prompt += f"Specialty / Type: {spec}\n"
    if size:
        prompt += f"Company size: {size}\n"

    prompt += (
        f"\nBased on these search results:\n{all_results}\n\n"
        f"Extract {limit_instruction} unique companies STRICTLY located in {location_for_search}.\n"
        f"Return ONLY a Python list of lists with exactly these 7 fields:\n"
        f'[["Company Name", "City", "Client Base", "Known Vendors", "Experience", "Website", "Company Size"]]\n\n'
        f"Rules:\n"
        f"- Client Base: Enterprise, Medium, Small, or Mixed\n"
        f"- Known Vendors: list ALL vendors/partners found, comma separated. Empty string if none.\n"
        f"- Experience: one sentence describing what the company does\n"
        f"- Website: EXACT URL from search results. Empty string if not found. NEVER guess or invent.\n"
        f"- Company Size: Small, Medium, or Large based on search results. Use Medium if unknown.\n"
        f"Return ONLY the raw Python list. No markdown. No explanation. "
        f"If no companies found return []."
    )

    return prompt

# ─────────────────────────────────────────────
# DISPLAY RESULTS AS CARDS
# ─────────────────────────────────────────────
def display_cards(user_profile):
    results = st.session_state.results
    if not results:
        return

    # Header
    col_count, col_export = st.columns([3, 1])
    with col_count:
        st.markdown(f"**{len(results)} {'company' if len(results) == 1 else 'companies'}** found")
    with col_export:
        if st.button("📥 Download Excel", use_container_width=True):
            export_to_excel(results)

    st.markdown("")

    # Cards in 2-column grid
    cols = st.columns(2, gap="medium")
    to_remove = None

    for i, row in enumerate(results):
        with cols[i % 2]:
            # Card HTML
            city    = row.get("City", "")
            size    = row.get("Company Size", "")
            cb      = row.get("Client Base", "")
            vendors = row.get("Known Vendors", "") or "—"
            about   = row.get("Experience", "")
            website = row.get("Website", "")
            name    = row.get("Company Name", "")

            badges = f'<span class="sr-badge sr-badge-city">{city}</span>' if city else ""
            if size: badges += f'<span class="sr-badge">{size}</span>'
            if cb:   badges += f'<span class="sr-badge">{cb} clients</span>'

            st.markdown(f"""
            <div class="sr-card">
                <div class="card-name">{name}</div>
                <div>{badges}</div>
                <hr class="card-divider"/>
                <div class="card-label">Known vendors</div>
                <div class="card-vendors">{vendors}</div>
                <div class="card-label">About</div>
                <div class="card-about">{about}</div>
            </div>
            """, unsafe_allow_html=True)

            # Action buttons
            btn_col1, btn_col2 = st.columns(2)
            with btn_col1:
                if website:
                    st.markdown(
                        f'<a href="{website}" target="_blank" style="font-size:12px; color:#1B6CA8; '
                        f'text-decoration:none; border:1px solid #BEDAF0; border-radius:4px; '
                        f'padding:4px 10px; background:#F0F7FC; display:inline-block;">🔗 Visit</a>',
                        unsafe_allow_html=True
                    )
            with btn_col2:
                if st.button("Exclude", key=f"exclude_{i}", use_container_width=True):
                    to_remove = i
                    exclude_from_suspects(name, website, user_profile)

    # Handle exclusion outside the loop
    if to_remove is not None:
        st.session_state.results.pop(to_remove)
        st.rerun()

# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────
def export_to_excel(results):
    try:
        df = pd.DataFrame(results)
        df.insert(0, "#", range(1, len(df) + 1))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for row_idx, row_data in enumerate(df.values.tolist(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        today    = datetime.now().strftime("%Y-%m-%d")
        location = st.session_state.search_location.replace(", ", "_").replace(" ", "_")
        filename = f"SalesRadar_{location}_{today}.xlsx"
        wb.save(filename)
        with open(filename, "rb") as f:
            st.download_button("📥 Download Excel", f, filename, use_container_width=True)
    except Exception as e:
        st.error("Export failed: " + str(e))

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
def show_app():
    anthropic_key = st.secrets["keys"]["ANTHROPIC_API_KEY"]
    serper_key    = st.secrets["keys"]["SERPER_API_KEY"]
    profile       = st.session_state.user_profile

    # ── HEADER ──────────────────────────────────
    col1, col2 = st.columns([4, 1])
    with col1:
        st.title("📡 Sales Radar")
    with col2:
        st.markdown(f"**{profile['name']}**")
        role_display = profile['role'].replace('_', ' ').title()
        st.markdown(f"{role_display}")
        if st.button("Sign Out", use_container_width=True):
            logout()

    st.divider()

    # ── LOAD ALL DATA ────────────────────────────
    cities_dict              = load_countries_and_cities()
    industry_names, id_map   = load_industries()

    if not cities_dict:
        st.error("Could not load configuration from database. Please try again.")
        return

    # ── COUNTRY LIST ─────────────────────────────
    if st.session_state.detected_country is None:
        st.session_state.detected_country = detect_user_country(sorted(list(cities_dict.keys())))
    user_country    = st.session_state.detected_country
    country_list    = build_country_list(cities_dict, user_country)
    country_options = ["All Countries"] + country_list

    # ── SIDEBAR ──────────────────────────────────
    with st.sidebar:
        st.markdown("#### 🎯 Search Criteria")

        # OBJECTIVE
        objective_options = [
            "Find companies to partner with",
            "Find companies to sell to",
            "Build a mailing list",
            "General research"
        ]
        objective = st.selectbox("Objective", objective_options, key="objective")

        # INDUSTRY
        industry_options = ["All Industries"] + industry_names if industry_names else ["All Industries"]
        industry = st.selectbox("Industry", industry_options, key="industry")

        # SPECIALTY — filtered by selected industry
        selected_industry_id = id_map.get(industry) if industry != "All Industries" else None
        specialty_list       = load_specialties(selected_industry_id)
        specialty_options    = ["All Specialties"] + specialty_list if specialty_list else ["All Specialties"]
        specialty = st.selectbox("Specialty", specialty_options, key="specialty")

        # SIZE
        size_options = ["Any Size", "Small (under 20)", "Medium (20-200)", "Large (200+)"]
        company_size = st.selectbox("Size", size_options, index=2, key="company_size")

        # COUNTRY
        if "country" not in st.session_state:
            default_idx = country_options.index(user_country) if user_country in country_options else 1
            st.session_state["country"] = country_options[default_idx]
        country = st.selectbox("Country", country_options, key="country")

        # CITY
        if country == "All Countries":
            city = "All Countries"
            st.caption("Searching worldwide.")
        else:
            city_options = ["All " + country] + cities_dict.get(country, [])
            if "city" in st.session_state and st.session_state["city"] not in city_options:
                st.session_state["city"] = city_options[0]
            city = st.selectbox("City", city_options, key="city")

        # LIMIT
        limit_on = st.checkbox("Limit results", value=False, key="limit_on")
        if limit_on:
            num_leads = st.number_input("Max results", min_value=1, value=10, step=5, key="num_leads")
        else:
            num_leads = 9999

        st.markdown("")
        search_button = st.button("🔍 Search", type="primary", use_container_width=True)

    # ── LOCATION LABELS ──────────────────────────
    if country == "All Countries":
        location_display    = "All Countries"
        location_for_search = "worldwide"
    elif city.startswith("All "):
        location_display    = country
        location_for_search = country
    else:
        location_display    = f"{city}, {country}"
        location_for_search = f"{city}, {country}"

    # ── SEARCH CONTEXT HEADER ────────────────────
    context_parts = [f"📍 {location_display}"]
    if industry != "All Industries":   context_parts.append(industry)
    if specialty != "All Specialties": context_parts.append(specialty)
    if company_size != "Any Size":     context_parts.append(company_size)
    st.markdown("### " + "  ·  ".join(context_parts))

    # ── SEARCH EXECUTION ─────────────────────────
    if search_button:
        # Clear previous results
        st.session_state.results       = []
        st.session_state.search_location = location_display

        client = anthropic.Anthropic(api_key=anthropic_key)

        # Get tenant name
        tenant_name = ""
        try:
            supabase = get_supabase_client()
            tenant   = supabase.table("tenants").select("name").eq("id", profile.get("tenant_id")).single().execute()
            if tenant.data:
                tenant_name = tenant.data.get("name", "")
        except:
            pass

        # BUILD AND RUN SEARCH QUERIES
        queries = build_search_queries(industry, specialty, company_size, location_for_search)

        with st.spinner(f"Searching for companies in {location_display}..."):
            all_results      = ""
            combined_url_map = {}
            for query in queries:
                snippets, url_map = search_web(query, serper_key)
                all_results += f"\nSearch: {query}\nResults:\n{snippets}\n"
                combined_url_map.update(url_map)

        # AI ANALYSIS
        with st.spinner("Analysing results..."):
            prompt = build_prompt(
                industry, specialty, company_size, objective,
                location_for_search, all_results, num_leads, tenant_name
            )

            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=8000,
                messages=[{"role": "user", "content": prompt}]
            )

            try:
                suspects = clean_and_parse(response.content[0].text)
            except Exception as e:
                st.error("Error parsing results: " + str(e))
                st.stop()

        if len(suspects) == 0:
            st.warning(f"No companies found in {location_display}. Try broadening your search criteria.")
        else:
            # BUILD DATAFRAME
            df = pd.DataFrame(suspects, columns=[
                "Company Name", "City", "Client Base",
                "Known Vendors", "Experience", "Website", "Company Size"
            ])

            # RESOLVE WEBSITE
            def resolve_website(row):
                claude_url = row.get("Website", "")
                if claude_url and is_directory_url(claude_url):
                    claude_url = ""
                domain = extract_domain(claude_url)
                if domain and domain in combined_url_map:
                    candidate = combined_url_map[domain]
                    if not is_directory_url(candidate):
                        return candidate
                return claude_url
            df["Website"] = df.apply(resolve_website, axis=1)

            # FILTER OWN COMPANY
            own_domains = load_own_domains(profile.get("tenant_id"))
            if own_domains:
                df = df[~df["Website"].apply(
                    lambda w: any(d in extract_domain(w) for d in own_domains)
                )].reset_index(drop=True)

            # APPLY LIMIT
            if num_leads < 9999:
                df = df.head(num_leads)

            # SAVE TO SUSPECTS
            with st.spinner("Saving to database..."):
                saved, skipped, excluded = save_to_suspects(
                    df,
                    country if country != "All Countries" else location_display,
                    objective,
                    industry,
                    specialty,
                    company_size,
                    profile
                )

            # SAVE SUMMARY
            parts = []
            if saved    > 0: parts.append(f"💾 {saved} new suspects saved")
            if skipped  > 0: parts.append(f"⏭️ {skipped} already existed")
            if excluded > 0: parts.append(f"🚫 {excluded} excluded — own company")

            if saved > 0:
                st.info(" | ".join(parts))
            elif skipped > 0:
                st.warning("⚠️ All results already exist in the database.")
            elif excluded > 0:
                st.warning("⚠️ All results were excluded — own company.")

            # STORE RESULTS IN SESSION STATE
            st.session_state.results = df.to_dict("records")

    # ── DISPLAY CARDS ───────────────────────────
    if st.session_state.results:
        display_cards(profile)
    elif not search_button:
        st.info("👈 Set your search criteria in the sidebar and click Search")

# ─────────────────────────────────────────────
# ROUTER
# ─────────────────────────────────────────────
if st.session_state.user is None:
    show_login()
else:
    show_app()
